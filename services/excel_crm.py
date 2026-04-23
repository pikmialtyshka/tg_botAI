from pathlib import Path
from threading import Lock
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class ExcelCRMService:
    FILE_PATH = Path('crm_leads.xlsx')
    SHEET_NAME = 'Contacts'
    SUMMARY_SHEET_NAME = 'Summary'
    HEADERS = ['ID', 'Username', 'Имя', 'Статус', 'Последнее сообщение пользователя', 'Последнее сообщение бота', 'Создан', 'Обновлён']

    _lock = Lock()

    def ensure_workbook(self):
        recreate = False
        if self.FILE_PATH.exists():
            try:
                wb = load_workbook(self.FILE_PATH)
                ws = wb[self.SHEET_NAME]
                headers = [ws.cell(row=1, column=i).value for i in range(1, len(self.HEADERS) + 1)]
                if headers != self.HEADERS or self.SUMMARY_SHEET_NAME not in wb.sheetnames:
                    recreate = True
            except Exception:
                recreate = True
        else:
            recreate = True

        if not recreate:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME
        self._setup_contacts_sheet(ws)
        self._setup_summary_sheet(wb)
        wb.save(self.FILE_PATH)

    def _setup_contacts_sheet(self, ws):
        ws.delete_rows(1, ws.max_row)
        ws.append(self.HEADERS)
        header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
        thin = Side(style='thin', color='D0D7DE')
        for idx, _ in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='center', horizontal='center')
        widths = {1: 8, 2: 26, 3: 24, 4: 16, 5: 28, 6: 28, 7: 22, 8: 22}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'

    def _setup_summary_sheet(self, wb):
        if self.SUMMARY_SHEET_NAME in wb.sheetnames:
            ws = wb[self.SUMMARY_SHEET_NAME]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(self.SUMMARY_SHEET_NAME)

        title_fill = PatternFill(fill_type='solid', fgColor='CFE2F3')
        label_fill = PatternFill(fill_type='solid', fgColor='EAF2F8')
        thin = Side(style='thin', color='D0D7DE')

        ws['A1'] = 'Сводка по CRM'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')

        rows = [
            ('Всего контактов', '=COUNTA(Contacts!A:A)-1'),
            ('Согласился', '=COUNTIF(Contacts!D:D,"согласился")'),
            ('Отказался', '=COUNTIF(Contacts!D:D,"отказался")'),
            ('Не ответил', '=COUNTIF(Contacts!D:D,"не ответил")'),
            ('Согласился (%)', '=IF(B3>0,B4/B3,0)'),
            ('Отказался (%)', '=IF(B3>0,B5/B3,0)'),
            ('Не ответил (%)', '=IF(B3>0,B6/B3,0)'),
            ('Обновлено', self._fmt_dt(datetime.now())),
        ]

        for row_idx, (label, value) in enumerate(rows, start=3):
            ws.cell(row=row_idx, column=1, value=label)
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            if label.endswith('(%)'):
                value_cell.number_format = '0.00%'
            ws.cell(row=row_idx, column=1).fill = label_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            for col in (1, 2):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 20
        ws.freeze_panes = 'A3'

    def _refresh_summary_sheet(self, wb):
        self._setup_summary_sheet(wb)

    def _load(self):
        self.ensure_workbook()
        wb = load_workbook(self.FILE_PATH)
        if self.SUMMARY_SHEET_NAME not in wb.sheetnames:
            self._setup_summary_sheet(wb)
        return wb, wb[self.SHEET_NAME]

    def upsert_contact(self, contact):
        if contact is None:
            return
        with self._lock:
            wb, ws = self._load()
            row_idx = None
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value or '') == str(contact.id or ''):
                    row_idx = row
                    break
            if row_idx is None:
                row_idx = ws.max_row + 1
            values = self._values(contact)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            self._refresh_summary_sheet(wb)
            wb.save(self.FILE_PATH)

    def sync_all_contacts(self, contacts):
        with self._lock:
            wb, ws = self._load()
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            for contact in contacts:
                ws.append(self._values(contact))
            self._refresh_summary_sheet(wb)
            wb.save(self.FILE_PATH)

    def _values(self, contact):
        name = ' '.join(part for part in [contact.first_name, contact.last_name] if part).strip() or None
        username = f'@{contact.username}' if contact.username else contact.user_id
        return [
            contact.id,
            username,
            name,
            contact.status,
            self._fmt_dt(contact.last_user_message_at),
            self._fmt_dt(contact.last_bot_message_at),
            self._fmt_dt(contact.created_at),
            self._fmt_dt(contact.updated_at),
        ]

    @staticmethod
    def _fmt_dt(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        return str(value)
